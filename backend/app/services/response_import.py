"""Response-sheet import (CSV / XLSX) into per-candidate answers.

Produces the ``ImportPreview`` shape the frontend renders before committing.
The sheet is matched by bank public id (``Q-1``) or by question position; cells
may hold a chosen option letter, a position number, or a right/wrong marker.
"""

import csv
import io
from typing import BinaryIO

from app.models import ExamPaper, Question
from openpyxl import load_workbook


class ImportError_(Exception):
    pass


def parse_sheet(db, paper: ExamPaper, file: BinaryIO, dry_run: bool) -> dict:
    # Decide the format from the content, not the extension.
    head = file.read(64)
    file.seek(0)
    if head.lstrip().startswith(b"PK"):
        rows = _read_xlsx(file)
    else:
        rows = _read_csv(file)
    if not rows:
        raise ImportError_("The file contained no data rows.")

    header = rows[0]
    body = rows[1:]
    return _build_preview(db, paper, header, body, dry_run)


def _read_csv(file: BinaryIO) -> list[list[str]]:
    text = file.read().decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if any(cell.strip() for cell in row)]


def _read_xlsx(file: BinaryIO) -> list[list[str]]:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row and any(cell is not None for cell in row):
            rows.append(["" if cell is None else str(cell).strip() for cell in row])
    wb.close()
    return rows


def _build_preview(db, paper: ExamPaper, header: list[str], body: list[list[str]], dry_run: bool) -> dict:
    questions_by_id = {str(q.question_id): q for q in paper.questions}
    questions_by_pos = {q.position: q for q in paper.questions}

    matched_by: list[str] = []
    lookups: dict[int, tuple[Question, int]] = {}

    # Header may carry a "bank id" column whose cells name the questions
    # (Q-1, Q-2...). If found, match by id.
    id_col = _find_column(header, ("bank id", "public id", "question id", "q id", "question"))
    if id_col is not None:
        matched_by.append("public_id")
        for idx, cell in enumerate(header):
            label = str(cell).strip().lower()
            for qid, q in questions_by_id.items():
                if f"q-{qid}" == label:
                    lookups[idx] = (q, q.position)
        if not lookups:
            matched_by = []

    # Otherwise fall back to positional matching: non-candidate columns are
    # question 1, 2, 3... in paper order.
    if not matched_by:
        matched_by.append("position")
        candidate_idxs = _candidate_columns(header)
        for idx, _cell in enumerate(header):
            if idx in candidate_idxs:
                continue
            q = questions_by_pos.get(len(lookups) + 1)
            if q:
                lookups[idx] = (q, q.position)
            else:
                lookups[idx] = (None, -1)

    candidate_keys = []
    answers = 0
    blank = 0
    duplicate_keys = set()
    rows_by_key: dict[str, list[list[str]]] = {}

    for row in body:
        if not any(cell for cell in row):
            continue
        key = _candidate_key(row)
        if key in rows_by_key:
            duplicate_keys.add(key)
        rows_by_key.setdefault(key, []).append(row)
        if key:
            candidate_keys.append(key)

    sample_candidates = list(dict.fromkeys(candidate_keys))[:10]
    warnings: list[str] = []
    if duplicate_keys:
        warnings.append(f"{len(duplicate_keys)} candidates appeared more than once; last row wins.")

    for key in duplicate_keys:
        rows_by_key[key] = [rows_by_key[key][-1]]

    answer_mode: str | None = None
    answered_questions: set[int] = set()
    for rows in rows_by_key.values():
        row = rows[0]
        for idx, (q, pos) in lookups.items():
            cell = row[idx] if idx < len(row) else ""
            if not cell.strip():
                blank += 1
                continue
            parsed, mode = _parse_answer(cell)
            if answer_mode is None:
                answer_mode = mode
            if parsed is not None and q is not None:
                answered_questions.add(pos)
                answers += 1
            else:
                blank += 1

    questions_matched = len({pos for _, pos in lookups.values() if pos != -1})
    used_columns = set(lookups.keys())
    unmatched_columns = [
        str(h)
        for h in header
        if h and h.strip() and all(idx not in used_columns for idx in range(len(header)) if header[idx] == h)
    ][:10]

    preview = {
        "dryRun": dry_run,
        "committed": not dry_run,
        "mode": "option" if answer_mode == "option" else "correct",
        "candidates": len(dict.fromkeys(candidate_keys)),
        "questionsMatched": questions_matched,
        "answers": answers,
        "matchedBy": matched_by,
        "unmatchedColumns": unmatched_columns,
        "questionsWithoutColumn": [str(pos) for pos in sorted(questions_by_pos) if pos not in answered_questions][:10],
        "duplicateCandidates": sorted(duplicate_keys)[:10],
        "blankCells": blank,
        "warnings": warnings,
        "sampleCandidates": sample_candidates,
    }

    if not dry_run:
        preview["responsesWritten"] = _commit(db, paper, rows_by_key, lookups)
        preview["responsesSuperseded"] = 0
        preview["candidatesCreated"] = 0
        preview["calibrated"] = questions_matched
    return preview


def _find_column(header: list[str], labels: tuple[str, ...]) -> int | None:
    for idx, cell in enumerate(header):
        if any(sub in str(cell).strip().lower() for sub in labels):
            return idx
    return None


def _candidate_columns(header: list[str]) -> set[int]:
    out: set[int] = set()
    for idx, cell in enumerate(header):
        label = str(cell).strip().lower()
        if label in ("candidate", "candidate id", "name", "roll no", "roll number", "id", "student", "student id"):
            out.add(idx)
        elif label.startswith("candidate"):
            out.add(idx)
    if not out and header:
        out = {0, 1}
    return out


def _candidate_key(row: list[str]) -> str:
    for cell in row[:2]:
        if cell and cell.strip():
            return cell.strip()
    return ""


def _parse_answer(cell: str) -> tuple[int | None, str]:
    value = cell.strip().lower()
    letter = {"a": 0, "b": 1, "c": 2, "d": 3, "e": 4}.get(value[:1])
    if letter is not None and value[:1].isalpha() and len(value) <= 3:
        return letter, "option"
    if value in {"0", "1"}:
        return int(value), "correct"
    try:
        n = int(value)
        if 0 <= n <= 4:
            return n, "option"
    except ValueError:
        pass
    return None, "correct"


def _commit(db, paper: ExamPaper, rows_by_key: dict, lookups: dict[int, tuple[Question, int]]) -> int:
    from app.models import StudentResponse

    written = 0
    for key, rows in rows_by_key.items():
        row = rows[0]
        for idx, (q, _pos) in lookups.items():
            cell = row[idx] if idx < len(row) else ""
            if not cell.strip():
                continue
            selected, mode = _parse_answer(cell)
            if selected is None:
                continue
            is_correct = selected == next((o.position for o in q.question.options if o.is_correct), -1)

            # Replace any existing row for this candidate+question (the sheet supersedes).
            existing = (
                db.query(StudentResponse)
                .filter(
                    StudentResponse.exam_paper_id == paper.id,
                    StudentResponse.question_id == q.question_id,
                    StudentResponse.candidate_key == key,
                )
                .first()
            )
            if existing:
                existing.selected_position = selected
                existing.is_correct = is_correct
            else:
                db.add(
                    StudentResponse(
                        exam_paper_id=paper.id,
                        question_id=q.question_id,
                        candidate_key=key,
                        selected_position=selected,
                        is_correct=is_correct,
                    )
                )
            written += 1
    db.commit()
    return written
